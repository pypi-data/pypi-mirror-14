from file_readers.file_reader import BaseFileReader
from file_readers.data_collector import BaseDataCollector
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter


class ExcelReader(BaseFileReader):

    def __init__(self, data_collector, debug=False):
        if not isinstance(data_collector, BaseDataCollector):
            raise AttributeError("Provided data collector is not a subclass of BaseDataCollector!")
        self.data_collector = data_collector
        self.debug = debug
        self.sheet_to_read = None

    def set_sheet_to_read(self, sheet_name):
        self.sheet_to_read = sheet_name

    @staticmethod
    def get_sheet_names(filename):
        work_book = load_workbook(filename)
        return work_book.get_sheet_names()

    def read_file(self, filename):
        work_book = load_workbook(filename)
        if self.sheet_to_read is None:
            self.sheet_to_read = work_book.get_active_sheet()
        sheet = work_book.get_sheet_by_name(self.sheet_to_read)
        rows = sheet.max_row
        columns = sheet.max_column
        min = 'A1'
        max_letter = get_column_letter(columns)
        max = max_letter + str(rows)

        for rowOfCellObjects in sheet[min:max]:
            self.data_collector.collect_data(rowOfCellObjects)
