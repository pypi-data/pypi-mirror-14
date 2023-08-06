#!/usr/bin/env python
# -*- coding: utf-8 -*-


"""phone2address CLI

Usage:
    p2a <phone>
    p2a -f <file>
"""

from bs4 import BeautifulSoup
from docopt import docopt
import openpyxl
import requests
import time


def save_data(data, file):
    data.save(file)
    return True


def load_data(file):
    data = []
    wb = openpyxl.load_workbook(file).active
    i = 2
    while (wb.cell(row=i, column=1).value != None):
        tmp = wb.cell(row=i, column=1).value
        if type(tmp) == type(""):
            data.append(tmp[:-2])
        else:
            data.append(str(int(tmp)))
        i += 1
    return data


def get_address(phone):
    url = 'http://www.ip138.com:8080/search.asp?action=mobile&mobile=%s' % (phone)
    r = requests.get(url)
    r.encoding = 'GBK'
    soup = BeautifulSoup(r.text, 'html.parser')
    try:
        data = soup.find_all('td')[6].text.split('\xa0')
    except IndexError:
        data = False
    print(data)
    return data


def process(file):
    data = load_data(file)
    wb = openpyxl.Workbook()
    ws = wb.active
    cur = 1
    for i in data:
        time.sleep(0.5)
        r = get_address(i)
        if r == False:
            ws.cell(row=cur, column=1).value = 'No a valid number'
        elif len(r) == 1:
            ws.cell(row=cur, column=2).value = data[0]
        else:
            ws.cell(row=cur, column=2).value = data[0]
            ws.cell(row=cur, column=3).value = data[1]
        cur += 1
    save_data(wb, 'processed.xlsx')


def main():
    arguments = docopt(__doc__, version='0.0.3')
    if arguments['-f']:
        process(arguments['<file>'])
    else:
        r = get_address(arguments['<phone>'])
        if r == False:
            print('请输入有效的手机号码')
        else:
            print('您的手机归属地为%s%s' % (r['province'], r['city']))


if __name__ == '__main__':
    main()
